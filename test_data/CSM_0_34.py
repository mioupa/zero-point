# 2024/02/28 alpha ver.0.7.0
# Created by Yuma Furukawa
# Please check the package below; xlrd

import tkinter as tk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from tkinter import scrolledtext
import numpy as np
from tkinter import filedialog
import pandas as pd
import math
import csv
import openpyxl
import os

#-------------------------グローバル変数宣言----------------------------------
# default test number
test_num = 1
# default zero point
zero = 0
# default scrollbar resolution
res = 1
# Area function
C0 = 21.2922
C1 = 2606.91
C2 = -42524.5
C3 = -81251.5
C4 = -19219
C5 = 79527.2
C6 = 156020
C7 = 204446

# C0 = 21.803
# C1 = 1670.13
# C2 = -25195.5
# C3 = -45516
# C4 = -9543.33
# C5 = 44676.6
# C6 = 88622.1
# C7 = 112672

# C0 = 25.6179
# C1 = -4433.46
# C2 = 61884.8
# C3 = 154574
# C4 = 68396.1
# C5 = -114111
# C6 = -265389
# C7 = -366045

# poisson ratio
poissonX = 0.18
# poissonDia = 0.07

#do not touch
x = []
y_load = []
y_HCS = []
y_mod = []
list4dep = []
list4lad = []
list4dep2 = []
list4dep3 = []
list4output = []
output_HCS = []
list4output2 = []
output_HCS2 = []
list4output3 = []
output_HCS3 = []
list4output4 = []
output_HCS4 = []
list4output5 = []
output_HCS5 = []
zero_stack = []
list4E = []
list4H = []
a = 0
b = 0
c = 0
dd = 0
x_line = []
y_line = []
x_line2 = []
y_line2 = []
x_line3 = []
y_line3 = []


#----------------------------ファイルを読み込み--------------------------------
filename = filedialog.askopenfilename(
    title = "ゼロ点補正するエクセルファイルを選択",
    filetypes = [ ("Excelブック", ".xls .xlsx") ], # ファイルフィルタ
    initialdir = "./" # 自分自身のディレクトリ
    )
basename = os.path.basename(filename)
#---------------------------------------------------Application-------------------------------------------------------
class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title('Zero Point Treatment')
        self.pack()
        # myfont = font.Font(root,family="Arial")
        self.create_widgets()
        # self.create_widgets_scrollbar()
        self.start_up()
        self.make_df()
        self.draw_plot()
        #アプリ起動時にここに並べた関数(子関数)を実施する親関数
        print(filename)

    def create_widgets(self):
        global text_out,res
        res = 1
        # self.control_af = tk.Frame(self.master,width=1500,height=30,borderwidth=1)
        # self.control_af.pack(side=tk.TOP, expand=True)
        self.canvas_frame = tk.Frame(self.master,width=1500,height=300,borderwidth=1)
        self.canvas_frame.pack(side=tk.TOP, expand=1)
        self.control_button = tk.Frame(self.master,width=1500,height=80,borderwidth=1)
        self.control_button.pack(side=tk.TOP, expand=1)
        self.control_zeropoint = tk.Frame(self.master,width=300,height=250,borderwidth=1)
        self.control_zeropoint.pack(side=tk.LEFT, expand=1)
        # self.control_res = tk.Frame(self.master,width=1200,height=80,borderwidth=1)
        # self.control_res.pack(side=tk.TOP, expand=1)
        self.control_frame = tk.Frame(self.master,width=1200,height=170,borderwidth=1)
        self.control_frame.pack(side=tk.TOP, expand=1)
        #フレームを作成して配置する

        self.canvas = FigureCanvasTkAgg(fig, self.canvas_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(side=tk.TOP)

        self.toolbar = NavigationToolbar2Tk(self.canvas, self.canvas_frame)
        self.toolbar.update()
        self.canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # self.setaf = tk.Button(self.control_af,
        #     text='SET AREA FUNCTION',
        #     command=lambda:[self.set_af(),self.make_df(),self.start_up(),self.draw_plot()]
        #     ).pack(side=tk.LEFT,anchor=tk.NW,fill='x')
        
        self.label_num = tk.Label(self.control_frame,text="number : ",font=("",10))
        self.label_num.pack(side=tk.LEFT,anchor="n")
        
        self.label_num = tk.Label(self.control_frame,text=test_num,font=("",10))
        self.label_num.pack(side=tk.LEFT,anchor="n")

        #display tip name
        wb = openpyxl.load_workbook(filename)
        tip= wb["Results"].cell(3,6).value
        self.label_tip = tk.Label(self.control_button,text='Tip name:',font=("",10)).pack(side=tk.LEFT,anchor="n")
        self.label_tip = tk.Label(self.control_button,text=tip,font=("",10)).pack(side=tk.LEFT,anchor=tk.W)

        self.button2 = tk.Button(self.control_button,
            text='NEXT TEST IF NO DATA',
            command=lambda:[self.stack(),self.go(),self.make_df(),self.start_up(),self.draw_plot()]
            )
        self.button2.pack(side=tk.LEFT,anchor=tk.CENTER,fill='x',expand=1)
        #次のテストへ進むボタン

        self.button1 = tk.Button(self.control_button,
            text='PREVIOUS TEST',
            command=lambda:[self.go_back(),self.make_df(),self.start_up(),self.draw_plot()]
            )
        self.button1.pack(side=tk.LEFT,anchor=tk.CENTER,fill='x',expand=1)
        #前のテストへ戻るボタン

        self.button3 = tk.Button(self.control_button,
            text='OUTPUT and NEXT TEST',
            command=lambda:[self.output(),self.stack(),self.go(),self.make_df(),self.start_up(),self.draw_plot()]
            )
        self.button3.pack(side=tk.LEFT,anchor=tk.CENTER,fill='x',expand=1)
        #出力ボタン
    
        # self.label_textout = tk.Label(self.control_frame,text='zero-treatment value',font=("",10)).pack(side=tk.LEFT)
        self.text_out=scrolledtext.ScrolledText(self.control_zeropoint,wrap='word',width=22,height=8)
        self.text_out.pack(side=tk.LEFT,anchor=tk.E,expand=1)
        #テキストボックスの作成

# set resolution button

        # self.coarse = tk.Button(self.control_res,
        #     text='Coarse',
        #     command=lambda:[self.set_res_coarse(),self.create_widgets_scrollbar(),self.draw_plot()]
        #     )
        # self.coarse.pack(side=tk.LEFT, anchor=tk.NW)

        # self.nomal = tk.Button(self.control_res,
        #     text='Nomal',
        #     command=lambda:[self.set_res_coarse(),self.create_widgets_scrollbar(),self.draw_plot()]
        #     )
        # self.nomal.pack(side=tk.LEFT, anchor=tk.NW)

        # self.fine = tk.Button(self.control_res,
        #     text='Fine',
        #     command=lambda:[self.set_res_coarse(),self.create_widgets_scrollbar(),self.draw_plot()]
        #     )
        # self.fine.pack(side=tk.LEFT, anchor=tk.NW)

    # def create_widgets_scrollbar(self):
        self.zero_w = tk.DoubleVar()
        self.zero_scale = tk.Scale(self.control_frame,
            variable=self.zero_w,
            length=1200,
            width=30,
            from_=-500,
            to=4000,
            resolution=res,
            orient=tk.HORIZONTAL,
            tickinterval=250,
            cursor='hand2',
            command=self.draw_plot)
        self.zero_scale.pack(side=tk.TOP,anchor=tk.SE)
        #スクロールバーの作成

    def start_up(self):
        self.zero_w.set(0)
        #アプリ起動時のzero-pointのデフォルトを設定

#-----------------------set the area function------------------------
    def set_af(self):
        C0 = self.area_c0_ent.get()
        C1 = self.area_c1_ent.get()
        C2 = self.area_c2_ent.get()
        C3 = self.area_c3_ent.get()
        C4 = self.area_c4_ent.get()
        C5 = self.area_c5_ent.get()
        C6 = self.area_c6_ent.get()
        C7 = self.area_c7_ent.get()
#-----------------------set the zero scale bar resolution------------------------
    def set_res_coarse(self):
        global res
        res = 10

    def set_res_nomal(self):
        global res
        res = 1

    def set_res_fine(self):
        global res
        res = 0.1

#-----------------------データフレームを作成---------------------------
    def make_df(self):
        global sheetname,sheetnames,test_num,df_load,df_HCS,df_mod,zero,afterT,list4dep
        wb = openpyxl.load_workbook(basename)
        sheetnames = wb.sheetnames
        print(sheetnames[len(sheetnames)-test_num - 1])
        # print("*")
        if str(sheetnames[len(sheetnames) - test_num - 1]) == 'Test {:03}'.format(test_num):
            df_load = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num),usecols=[1,2],skiprows=[0])
            df_HCS = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num),usecols=[1,4],skiprows=[0])
            df_mod = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num),usecols=[1,6],skiprows=[0])
            #データフレームを作成

            for line in range(0,df_load.shape[0]):
                # if np.isnan(df_load.iat[line,0]):
                #     break
                x.append(df_load.iat[line,0])
                y_load.append(df_load.iat[line,1])
                y_HCS.append(df_HCS.iat[line,1])
                y_mod.append(df_mod.iat[line,1])
                list4dep.append(df_HCS.iat[line,0]) # df_HCS.iat[line,0] is depth
            #データフレームに値を格納

            for k in range(1,3000000,10000):
                x_line.append(2*k)
                y_line.append(1000*k)
                x_line2.append(k)
                y_line2.append(1000*k)
                x_line3.append(k/2)
                y_line3.append(1000*k)
            #ゼロ点補正時のガイドラインを作成（黄色の線）

            line_low.set_xdata(x_line)
            line_low.set_ydata(y_line)
            line_mid.set_xdata(x_line2)
            line_mid.set_ydata(y_line2)
            line_high.set_xdata(x_line3)
            line_high.set_ydata(y_line3)
            #ガイドラインの値を引き渡し
        if str(sheetnames[len(sheetnames) - test_num - 1]) != 'Test {:03}'.format(test_num):
            df_load = pd.DataFrame()
            df_HCS = pd.DataFrame()
            df_mod = pd.DataFrame()
            print(zero)
            print('Test {:03} does not exist'.format(test_num))
            pass            

            
#!!out of indexでエラー吐く!!-->breakのif文の条件式に、Noneは使えない！(dfの空白にはNoneではなくNanという文字列が入ってる)

    def go(self):
        global test_num
        test_num += 1
        print('\n')
        print("*************************go to Test {:03}".format(test_num))
        self.label_num["text"] = test_num
        #次のテストへのボタンを押したときに呼び出される関数

    def go_back(self):
        global test_num
        test_num -= 1
        self.label_num["text"] = test_num
        #前のテストへのボタンを押したときに呼び出される関数

    def draw_plot(self, event=None):
        global h_l,h_H,h_m,x,y_load,y_HCS,y_mod,co,text_out,line_low,line_mid,line_high,list4dep,list4lad,afterT,afterL,zero,answerE,answerdep,list4E,list4dep2,a,b,flag,dd
        zero = self.zero_w.get()
        #スクロールバーの値を取得

        afterT = 0
        afterL = 0
        answerE = 0
        answerdep = 0
        flag = 0
        #変数を定義
        
        for line in range(0,df_load.shape[0]):
            if np.isnan(df_load.iat[line,0]):
                break
            x.append(df_load.iat[line,0])
            y_load.append(df_load.iat[line,1])
            y_HCS.append(df_HCS.iat[line,1])
            y_mod.append(df_mod.iat[line,1])
        # #実験データを配列に格納

        for l in range(0,df_HCS.shape[0]):
            if np.isnan(df_HCS.iat[l,0]):
                break
            afterT = df_HCS.iat[l,0] - zero
            list4dep.append(afterT)
            # dd = list4dep[l-1]
            # a = C0*(dd)**2 + C1*dd + C2*(dd**(1/2)) + C3*(dd**(1/4)) + C4*(dd**(1/8)) + C5*(dd**(1/16)) + C6*(dd**(1/32)) + C7*(dd**(1/64))
            hc = list4dep[l-1] - 0.75 * 1.0e6 * df_load.iat[l,1] / df_HCS.iat[l,1] # contact depth
            aa = C0*(hc)**2+C1*hc+C2*(hc**(1/2))+C3*(hc**(1/4))+C4*(hc**(1/8))+C5*(hc**(1/16))+C6*(hc**(1/32))+C7*(hc**(1/64)) # area function
            b = df_HCS.iat[l,1]
            # afterE = 1/(((1-0.0784)/(y_mod[l-1])) + ((1-(poissionDia)**2)/1140))
            if aa < 0:
                continue
            # answerE = (b*(math.sqrt(math.pi)))/(2.068*(math.sqrt(a)))
            # (1 - ThisWorkbook.Sheets(1).Cells(1, 4) * ThisWorkbook.Sheets(1).Cells(1, 4)) * (1 / ((2 * Excel.Application.WorksheetFunction.Power(aa, 0.5) / (.Range("c" & j) * Excel.Application.WorksheetFunction.Power(WorksheetFunction.Pi(), 0.5))) - ((1 - 0.07 * 0.07) / 1141)))
            answerE = (1 - poissonX*poissonX)*(1 / ((2 * np.sqrt(aa) / (b * np.sqrt(np.pi)) - ((1 - 0.07*0.07)/1141))))

            list4E.append(answerE)
        #ゼロ点補正後の弾性率を計算
        
        for n in range(0,len(list4E)):
            answerdep = list4dep[n]
            list4dep2.append(answerdep)

        h_l.set_xdata(x)
        h_l.set_ydata(y_load)
        h_H.set_xdata(list4dep)
        h_H.set_ydata(y_HCS)
        h_m.set_xdata(list4dep2)
        h_m.set_ydata(list4E)
        #データフレームの中の値を引き継ぐ

        x = []
        y_load = []
        y_HCS = []
        y_mod = []
        list4dep = []
        list4dep2 = []
        list4E = []
        a = 0
        b = 0
        dd = 0
        #データフレームを空集合として再定義してクリアに

        self.canvas.draw()
        #グラフを更新

    def output(self):
        global zero,text_out,flag,test_num,g_stack,list4output,output_HCS,g_stack2,list4output2,output_HCS2,g_stack3,list4output3,output_HCS3,g_stack4,list4output4,output_HCS4,g_stack5,list4output5,output_HCS5, x, list4dep, list4E, list4H, df_load
        flag = 0
        self.text_out.insert(tk.END,zero)
        self.text_out.insert(tk.END,'\n')
        #テキストを出力したときに実行する

        #HCSに対してゼロ点補正を行う
        for line in range(0,df_HCS.shape[0]):
            if np.isnan(df_HCS.iat[line,0]):
                break
            output_afterT = df_HCS.iat[line,0] - zero
            afterL = df_load.iat[line,1]
            list4output.append(output_afterT)
            output_HCS.append(df_HCS.iat[line,1])
            #
            list4dep.append(output_afterT)
            # print("***************************")
            # print("({}) - 0.75 * 1.0e6 * ({}) / ({})".format(output_afterT, df_load.iat[line,1], output_HCS[line]))
            # print("= ({}) - ({})".format(output_afterT, 0.75 * 1.0e6 * df_load.iat[line,1] / output_HCS[line]))
            # dd = list4dep[line-1]
            hc = output_afterT - 0.75 * 1.0e6 * afterL / output_HCS[line] # contact depth
            # print(hc)
            aa = C0*(hc)**2+C1*hc+C2*(hc**(1/2))+C3*(hc**(1/4))+C4*(hc**(1/8))+C5*(hc**(1/16))+C6*(hc**(1/32))+C7*(hc**(1/64)) # area function
            b = df_HCS.iat[line,1]
            # afterE = 1/(((1-0.0784)/(y_mod[l-1])) + ((1-(poissionDia)**2)/1140))
            ###
            if aa < 0: # このif文はanswerの計算よりも絶対前に書くこと
                continue

            # indentation elastic modulus (GPa)
            # answerE = (b*(math.sqrt(math.pi)))/(2.068*(math.sqrt(a)))
            answerE = (1 - poissonX*poissonX)*(1 / ((2 * 1.034 * np.sqrt(aa) / (b * np.sqrt(np.pi)) - ((1 - 0.07*0.07)/1141))))
            # print(answerE)
            list4E.append(answerE)

            # indentation hardness (GPa)
            # list4lad.append(afterL)
            # hc = list4dep[line-1] - 0.75 * 1.0e6 * afterL / output_HCS[line]
            # aa = C0*(hc)**2+C1*hc+C2*(hc**(1/2))+C3*(hc**(1/4))+C4*(hc**(1/8))+C5*(hc**(1/16))+C6*(hc**(1/32))+C7*(hc**(1/64))
            if aa < 0:
                continue
            answerH = afterL / aa * 1.0e6
            list4H.append(answerH)

        # for i in range(0,len(list4E)):
        #     answerdep = list4dep[i]
        #     list4dep2.append(answerdep)
        
        # for i in range(0, len(list4H)):
        #     answerdep = list4dep[i]
        #     list4dep3.append(answerdep)
        # print("*********************************")
        # list4dep2 = list4dep[0:len(list4E)]
        list4dep3 = list4dep[0:len(list4H)]

            #ゼロ点補正後の弾性率を計算

# kokomade atteru
        fileeeeename = 'zero-point_correction_' + str(test_num) + '.csv'
        Header = ["Length for revising zero point (nm)", "Depth (nm)", "Harmonic contact stiffness (N/m)", "Modulus (GPa)", "Nanoindentation hardness (GPa)", "1/depth (1/nm)", "Hardness^2 (GPa^2)", "Bulk equivalent hardness (GPa)", "Characteristic length of ISE (nm)"] # header of output data
        # print('*******************************')

        # Nix-Gao
        # (indentation hardness)^2 - hの関係から最小2乗法を用いて線形近似し、fitting parameterを求める
        th = 200.0 # threshold depth for simple regression (nm)
        X = ([1 / list4dep3[i] for i in range(len(list4dep3)) if (list4dep3[i]>th) & (list4H[i]>0.0)]) # depth after zero treatment
        Y = ([list4H[i] * list4H[i] for i in range(len(list4H)) if (list4dep3[i]>th) & (list4H[i]>0.0)]) # (indentation hardness)^2
        print(len(X), len(Y))
        # print(Y)
        
        # fitting: y=sx+t
        C = np.cov(X, Y) # 共分散行列の計算
        s = C[0,1] / C[0,0] # 近似直線の傾き
        t = np.mean(Y) - s*np.mean(X) # 近似直線の切片
        BeqH = np.sqrt(t) # Bulk equivalent hardness
        hstar = s / t # hstar
        
        # output: Length for revising zero point (nm), Depth (nm), HCS (N/m), Modulus (GPa), H_IT (GPa), 1/depth (1/nm), Hardness^2 (GPa^2), BeqH (GPa), hstar (nm)
        data_output = [[zero], list4output, output_HCS, list4E, list4H, X, Y, [BeqH], [hstar]]
        # print(len(row) for row in data_output)

        # Noneで不足分を補ったリストを作成
        padded_data = [row + [None] * (len(list4output) - len(row)) for row in data_output]
        # NumPyの2D配列に変換
        array_with_none = np.array(padded_data, dtype=object)

        with open(fileeeeename,'w',newline="") as f:
            writer = csv.writer(f)
            writer.writerow(Header)
            writer.writerows(array_with_none.T)
        
    #output_HCSとlist4Eをcsvに出力 

        #ゼロ点補正を行ったテスト結果を5つまで保存する
    def stack(self):
        global zero,text_out,flag,test_num,g_stack,list4output,output_HCS,g_stack2,list4output2,output_HCS2,g_stack3,list4output3,output_HCS3,g_stack4,list4output4,output_HCS4,g_stack5,list4output5,output_HCS5, x, list4dep, list4E, list4H, df_load
        # 1つ前のグラフ
        # print(str(sheetnames[- test_num - 1 + 1]))
        if test_num >=2:
            print('-1'+str(sheetnames[- test_num]) + ' : Test' + str(test_num-1))
            # Test00Xが存在する場合に行う処理
            if str(sheetnames[- test_num]) == 'Test {:03}'.format(test_num-1):
                predf_HCS = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num-1),usecols=[1,4],skiprows=[0])
                for line in range(1,1500):
                    if np.isnan(predf_HCS.iat[line,0]):
                        break
                    output_afterT2 = predf_HCS.iat[line,0] - zero_stack[test_num-2]
                    output_HCS2.append(predf_HCS.iat[line,1])
                    list4output2.append(output_afterT2)
            # Test00Xが存在しない場合に行う処理
            if  str(sheetnames[- test_num]) != 'Test {:03}'.format(test_num-1):
                output_afterT2 = 0
                output_HCS2.append(0)
                list4output2.append(output_afterT2)
            # Test00Xの存在に関わらず行う処理
            # output_HCS2.append(predf_HCS.iat[line,1])

        # 2つ前のグラフ
        # print(str(sheetnames[- test_num - 1 + 2]))
        if test_num >=3:
            print('-2' + str(sheetnames[- test_num + 1]) + ' : Test' + str(test_num-2))
            # Test00Xが存在する場合に行う処理
            if str(sheetnames[- test_num + 1]) == 'Test {:03}'.format(test_num-2):
                predf_HCS2 = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num-2),usecols=[1,4],skiprows=[0])
                for line in range(1,1500):
                    if np.isnan(predf_HCS2.iat[line,0]):
                        break
                    output_afterT3 = predf_HCS2.iat[line,0] - zero_stack[test_num-3]
                    output_HCS3.append(predf_HCS2.iat[line,1])
                    list4output3.append(output_afterT3) 
            # Test00Xが存在しない場合に行う処理
            if  str(sheetnames[- test_num + 1]) != 'Test {:03}'.format(test_num-2):
                output_afterT3 = 0
                output_HCS3.append(0)
                list4output3.append(output_afterT3) 
            # Test00Xの存在に関わらず行う処理
            # output_HCS3.append(predf_HCS2.iat[line,1])

        # 3つ前のグラフ
        # print(str(sheetnames[- test_num - 1 + 3]))
        if test_num >=4:
            print('-3'+str(sheetnames[- test_num + 2]) + ' : Test' + str(test_num-3))
            # Test00Xが存在する場合に行う処理
            if str(sheetnames[- test_num + 2]) == 'Test {:03}'.format(test_num-3):
                predf_HCS3 = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num-3),usecols=[1,4],skiprows=[0])
                for line in range(1,1500):
                    if np.isnan(predf_HCS3.iat[line,0]):
                        break
                    output_afterT4 = predf_HCS3.iat[line,0] - zero_stack[test_num-4]
                    output_HCS4.append(predf_HCS3.iat[line,1])
                    list4output4.append(output_afterT4) 
            # Test00Xが存在しない場合に行う処理
            if  str(sheetnames[- test_num + 2]) != 'Test {:03}'.format(test_num-3):
                output_afterT4 = 0
                output_HCS4.append(0)
                list4output4.append(output_afterT4) 
            # Test00Xの存在に関わらず行う処理
            # output_HCS4.append(predf_HCS3.iat[line,1])

        # 4つ前のグラフ      
        # print(str(sheetnames[- test_num - 1 + 4]))
        if test_num >=5:
            print('-4'+str(sheetnames[- test_num + 3]) + ' : Test' + str(test_num-4))
            # Test00Xが存在する場合に行う処理
            if str(sheetnames[- test_num + 3]) == 'Test {:03}'.format(test_num-4):
                predf_HCS4 = pd.read_excel(filename,sheet_name='Test {:03}'.format(test_num-4),usecols=[1,4],skiprows=[0])
                for line in range(1,1500):
                    if np.isnan(predf_HCS4.iat[line,0]):
                        break
                    output_afterT5 = predf_HCS4.iat[line,0] - zero_stack[test_num-5]
                    output_HCS5.append(predf_HCS4.iat[line,1])
                    list4output5.append(output_afterT5)
            # Test00Xが存在しない場合に行う処理
            if  str(sheetnames[- test_num + 3]) != 'Test {:03}'.format(test_num-4):
                output_afterT5 = 0
                output_HCS5.append(0)
                list4output5.append(output_afterT5)
            # Test00Xの存在に関わらず行う処理
            # output_HCS5.append(predf_HCS4.iat[line,1])
        print(zero)
        zero_stack.append(zero)
        print(len(zero_stack))
        g_stack.set_xdata(list4output)
        g_stack.set_ydata(output_HCS)
        g_stack2.set_xdata(list4output2)
        g_stack2.set_ydata(output_HCS2)
        g_stack3.set_xdata(list4output3)
        g_stack3.set_ydata(output_HCS3)
        g_stack4.set_xdata(list4output4)
        g_stack4.set_ydata(output_HCS4)
        g_stack5.set_xdata(list4output5)
        g_stack5.set_ydata(output_HCS5)
        # ゼロ点補正を行ったテスト結果の値を引き渡し

        list4output = []
        output_HCS = []
        list4output2 = []
        output_HCS2 = []
        list4output3 = []
        output_HCS3 = []
        list4output4 = []
        output_HCS4 = []
        list4output5 = []
        output_HCS5 = []
        list4E = []
        list4H = []
        list4dep = []
        list4dep2 = []
        list4dep3 = []
        d = []
        h = []
        m = []
        n = []
        p = []
        q = []
        o = []
        #リストを再定義して空集合にする

#------------------------------------------------------------------------------------------------------

#---------------------------グラフを描写---------------------------
fig = Figure(figsize=(15,5))

#左のグラフ
ax = fig.add_subplot(131)
ax.set_xlim(0,3500)
ax.set_ylim(0,300)
ax.set_ylabel("load(mN)")
h_l, = ax.plot([],[],color='teal')
#グラフの設定。引数を空欄にすることで、動的にグラフを変更する

#真ん中のグラフ
ax2 = fig.add_subplot(132)
ax2.set_xlim(0,3500)
ax2.set_ylim(0,3000000)
ax2.set_xlabel("Depth(nm)")
ax2.set_ylabel("Harmonic contact stiffness(N/m)")
h_H, = ax2.plot([],[],color='tomato')
line_low, = ax2.plot([],[],color='y')
line_mid, = ax2.plot([],[],color='y')
line_high, = ax2.plot([],[],color='y')
g_stack, = ax2.plot([],[],color='dodgerblue', alpha=0.7)
g_stack2, = ax2.plot([],[],color='dodgerblue', alpha=0.6)
g_stack3, = ax2.plot([],[],color='dodgerblue', alpha=0.5)
g_stack4, = ax2.plot([],[],color='dodgerblue', alpha=0.4)
g_stack5, = ax2.plot([],[],color='dodgerblue', alpha=0.3)

#右のグラフ
ax3 = fig.add_subplot(133)
ax3.set_xlim(0,3500)
ax3.set_ylim(0,1000)
ax3.set_ylabel("Modulus(GPa)")
h_m, = ax3.plot([],[],color='teal')

root = tk.Tk()
root.geometry("1500x700")
#アプリケーションの初期ウィンドウのサイズ
app = Application(master=root)
app.mainloop()